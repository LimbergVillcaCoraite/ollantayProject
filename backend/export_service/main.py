#!/usr/bin/env python3
"""
Export Service - Generate PDF and Excel reports
Provides endpoints to export data from various modules (compras, ventas, prestamos, asistencia, etc.)
"""

from fastapi import FastAPI, HTTPException, Depends, Query
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List, Literal
import mysql.connector
import os
import jwt
from datetime import datetime, timedelta
import io

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
except ImportError:
    print("WARNING: reportlab not installed. PDF export will not work.")
    print("Install with: pip install reportlab")

# Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("WARNING: openpyxl not installed. Excel export will not work.")
    print("Install with: pip install openpyxl")

app = FastAPI(title="Export Service", version="1.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database connection
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'mysql'),
    'port': int(os.getenv('DB_PORT', 3306)),
    'user': os.getenv('DB_USER', 'ollantay_user'),
    'password': os.getenv('DB_PASSWORD', 'ollantay_password'),
    'database': os.getenv('DB_NAME', 'SystemaOllantay')
}

JWT_SECRET = os.getenv('JWT_SECRET', 'ollantay-super-secret-jwt-key-change-in-production')
JWT_ALGORITHM = 'HS256'

def get_db():
    conn = mysql.connector.connect(**DB_CONFIG)
    try:
        yield conn
    finally:
        conn.close()

def verify_token(cookie_str: str):
    """Extract and verify JWT from cookie string"""
    if not cookie_str:
        raise HTTPException(401, "No token provided")
    
    # Parse cookie string to extract token
    token = None
    for part in cookie_str.split(';'):
        part = part.strip()
        if part.startswith('ollantay_token='):
            token = part.split('=', 1)[1]
            break
    
    if not token:
        raise HTTPException(401, "Token not found in cookies")
    
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

# Dependency to get user from token
async def get_current_user(cookie: Optional[str] = None):
    if not cookie:
        raise HTTPException(401, "Authentication required")
    return verify_token(cookie)

# ============= COMPRAS EXPORT =============

@app.get("/export/compras")
async def export_compras(
    format: Literal["pdf", "excel"] = Query("pdf"),
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    id_proveedor: Optional[int] = None,
    conn=Depends(get_db),
    user=Depends(get_current_user)
):
    """Export compras report in PDF or Excel format"""
    company_id = user.get('company_id')
    if not company_id:
        raise HTTPException(400, "company_id required")
    
    cursor = conn.cursor(dictionary=True)
    
    # Build query
    query = """
        SELECT 
            c.id_compra,
            c.fechaCompra,
            c.numeroFactura,
            c.montoTotal,
            p.nombres_proveedor,
            p.apellido_paterno_proveedor,
            tp.nombre_tipo_pago,
            c.estado_pago
        FROM compra_O c
        LEFT JOIN proveedor_O p ON c.idProveedor = p.id_proveedor AND p.idEmpresa = c.idEmpresa
        LEFT JOIN tipo_pago_O tp ON c.idTipoPago = tp.id_tipo_pago AND tp.idEmpresa = c.idEmpresa
        WHERE c.idEmpresa = %s
    """
    params = [company_id]
    
    if fecha_inicio:
        query += " AND c.fechaCompra >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND c.fechaCompra <= %s"
        params.append(fecha_fin)
    if id_proveedor:
        query += " AND c.idProveedor = %s"
        params.append(id_proveedor)
    
    query += " ORDER BY c.fechaCompra DESC LIMIT 1000"
    
    cursor.execute(query, params)
    compras = cursor.fetchall()
    cursor.close()
    
    if not compras:
        raise HTTPException(404, "No data found")
    
    if format == "pdf":
        return generate_compras_pdf(compras, fecha_inicio, fecha_fin)
    else:
        return generate_compras_excel(compras, fecha_inicio, fecha_fin)

def generate_compras_pdf(compras, fecha_inicio, fecha_fin):
    """Generate PDF report for compras"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Reporte de Compras", title_style))
    
    # Date range
    if fecha_inicio or fecha_fin:
        date_text = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Fin'}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Summary
    total_compras = len(compras)
    total_monto = sum(float(c['montoTotal'] or 0) for c in compras)
    summary_text = f"Total de compras: {total_compras} | Monto total: Bs {total_monto:,.2f}"
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Fecha', 'Factura', 'Proveedor', 'Tipo Pago', 'Monto', 'Estado']]
    for c in compras:
        proveedor = f"{c['nombres_proveedor'] or ''} {c['apellido_paterno_proveedor'] or ''}".strip() or 'N/A'
        estado = 'Pagado' if c['estado_pago'] == 'pagado' else 'Pendiente' if c['estado_pago'] == 'pendiente' else 'Parcial'
        data.append([
            c['fechaCompra'].strftime('%d/%m/%Y') if c['fechaCompra'] else '-',
            c['numeroFactura'] or '-',
            proveedor[:25],
            c['nombre_tipo_pago'] or '-',
            f"Bs {float(c['montoTotal'] or 0):,.2f}",
            estado
        ])
    
    # Create table
    table = Table(data, colWidths=[1*inch, 1*inch, 1.8*inch, 1*inch, 1.2*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=compras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
    )

def generate_compras_excel(compras, fecha_inicio, fecha_fin):
    """Generate Excel report for compras"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Reporte de Compras"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Date range
    if fecha_inicio or fecha_fin:
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Fin'}"
        ws['A2'].alignment = Alignment(horizontal="center")
    
    # Summary
    total_compras = len(compras)
    total_monto = sum(float(c['montoTotal'] or 0) for c in compras)
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Total de compras: {total_compras} | Monto total: Bs {total_monto:,.2f}"
    ws['A3'].alignment = Alignment(horizontal="center")
    ws['A3'].font = Font(bold=True)
    
    # Headers
    headers = ['Fecha', 'Factura', 'Proveedor', 'Tipo Pago', 'Monto (Bs)', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    for row_num, c in enumerate(compras, 6):
        proveedor = f"{c['nombres_proveedor'] or ''} {c['apellido_paterno_proveedor'] or ''}".strip() or 'N/A'
        estado = 'Pagado' if c['estado_pago'] == 'pagado' else 'Pendiente' if c['estado_pago'] == 'pendiente' else 'Parcial'
        
        ws.cell(row=row_num, column=1, value=c['fechaCompra'].strftime('%d/%m/%Y') if c['fechaCompra'] else '-').border = border
        ws.cell(row=row_num, column=2, value=c['numeroFactura'] or '-').border = border
        ws.cell(row=row_num, column=3, value=proveedor).border = border
        ws.cell(row=row_num, column=4, value=c['nombre_tipo_pago'] or '-').border = border
        ws.cell(row=row_num, column=5, value=float(c['montoTotal'] or 0)).border = border
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6, value=estado).border = border
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=compras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

# ============= PRESTAMOS EXPORT =============

@app.get("/export/prestamos")
async def export_prestamos(
    format: Literal["pdf", "excel"] = Query("pdf"),
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    estado_prestamo: Optional[str] = None,
    conn=Depends(get_db),
    user=Depends(get_current_user)
):
    """Export prestamos report"""
    company_id = user.get('company_id')
    if not company_id:
        raise HTTPException(400, "company_id required")
    
    cursor = conn.cursor(dictionary=True)
    
    query = """
        SELECT 
            pr.id_prestamo,
            pr.fecha_prestamo,
            pr.monto_total,
            pr.interes,
            pr.estado_prestamo,
            pr.saldo_pendiente,
            p.nombres_persona,
            p.apellido_paternoPersona,
            p.ci_persona
        FROM prestamo_O pr
        LEFT JOIN persona_O p ON pr.id_persona = p.id_persona AND p.idEmpresa = pr.idEmpresa
        WHERE pr.idEmpresa = %s
    """
    params = [company_id]
    
    if fecha_inicio:
        query += " AND pr.fecha_prestamo >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND pr.fecha_prestamo <= %s"
        params.append(fecha_fin)
    if estado_prestamo:
        query += " AND pr.estado_prestamo = %s"
        params.append(estado_prestamo)
    
    query += " ORDER BY pr.fecha_prestamo DESC LIMIT 1000"
    
    cursor.execute(query, params)
    prestamos = cursor.fetchall()
    cursor.close()
    
    if not prestamos:
        raise HTTPException(404, "No data found")
    
    if format == "pdf":
        return generate_prestamos_pdf(prestamos, fecha_inicio, fecha_fin)
    else:
        return generate_prestamos_excel(prestamos, fecha_inicio, fecha_fin)

def generate_prestamos_pdf(prestamos, fecha_inicio, fecha_fin):
    """Generate PDF report for prestamos"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1f2937'), spaceAfter=30, alignment=TA_CENTER)
    elements.append(Paragraph("Reporte de Préstamos", title_style))
    
    if fecha_inicio or fecha_fin:
        date_text = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Fin'}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    total_prestamos = len(prestamos)
    total_monto = sum(float(p['monto_total'] or 0) for p in prestamos)
    total_pendiente = sum(float(p['saldo_pendiente'] or 0) for p in prestamos)
    summary = f"Total préstamos: {total_prestamos} | Monto total: Bs {total_monto:,.2f} | Pendiente: Bs {total_pendiente:,.2f}"
    elements.append(Paragraph(summary, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    data = [['Fecha', 'Cliente', 'CI', 'Monto', 'Interés', 'Saldo', 'Estado']]
    for p in prestamos:
        cliente = f"{p['nombres_persona'] or ''} {p['apellido_paternoPersona'] or ''}".strip() or 'N/A'
        data.append([
            p['fecha_prestamo'].strftime('%d/%m/%Y') if p['fecha_prestamo'] else '-',
            cliente[:20],
            p['ci_persona'] or '-',
            f"Bs {float(p['monto_total'] or 0):,.2f}",
            f"{float(p['interes'] or 0):.1f}%",
            f"Bs {float(p['saldo_pendiente'] or 0):,.2f}",
            p['estado_prestamo'] or '-'
        ])
    
    table = Table(data, colWidths=[1*inch, 1.3*inch, 0.9*inch, 1*inch, 0.8*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=prestamos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
    )

def generate_prestamos_excel(prestamos, fecha_inicio, fecha_fin):
    """Generate Excel report for prestamos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "Reporte de Préstamos"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    if fecha_inicio or fecha_fin:
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Fin'}"
        ws['A2'].alignment = Alignment(horizontal="center")
    
    total_prestamos = len(prestamos)
    total_monto = sum(float(p['monto_total'] or 0) for p in prestamos)
    total_pendiente = sum(float(p['saldo_pendiente'] or 0) for p in prestamos)
    ws.merge_cells('A3:G3')
    ws['A3'] = f"Total: {total_prestamos} | Monto total: Bs {total_monto:,.2f} | Pendiente: Bs {total_pendiente:,.2f}"
    ws['A3'].alignment = Alignment(horizontal="center")
    ws['A3'].font = Font(bold=True)
    
    headers = ['Fecha', 'Cliente', 'CI', 'Monto (Bs)', 'Interés (%)', 'Saldo (Bs)', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row_num, p in enumerate(prestamos, 6):
        cliente = f"{p['nombres_persona'] or ''} {p['apellido_paternoPersona'] or ''}".strip() or 'N/A'
        ws.cell(row=row_num, column=1, value=p['fecha_prestamo'].strftime('%d/%m/%Y') if p['fecha_prestamo'] else '-').border = border
        ws.cell(row=row_num, column=2, value=cliente).border = border
        ws.cell(row=row_num, column=3, value=p['ci_persona'] or '-').border = border
        ws.cell(row=row_num, column=4, value=float(p['monto_total'] or 0)).border = border
        ws.cell(row=row_num, column=4).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=float(p['interes'] or 0)).border = border
        ws.cell(row=row_num, column=5).number_format = '0.0'
        ws.cell(row=row_num, column=6, value=float(p['saldo_pendiente'] or 0)).border = border
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'
        ws.cell(row=row_num, column=7, value=p['estado_prestamo'] or '-').border = border
    
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=prestamos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

# ============= ASISTENCIA EXPORT =============

@app.get("/export/asistencia")
async def export_asistencia(
    format: Literal["pdf", "excel"] = Query("pdf"),
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    id_persona: Optional[int] = None,
    conn=Depends(get_db),
    user=Depends(get_current_user)
):
    """Export asistencia report"""
    company_id = user.get('company_id')
    if not company_id:
        raise HTTPException(400, "company_id required")
    
    cursor = conn.cursor(dictionary=True)
    
    query = """
        SELECT 
            a.id_asistencia,
            a.fecha,
            a.hora_entrada,
            a.hora_salida,
            p.nombres_persona,
            p.apellido_paternoPersona,
            p.ci_persona,
            e.cargo
        FROM asistencia_O a
        LEFT JOIN persona_O p ON a.id_persona = p.id_persona AND p.idEmpresa = a.idEmpresa
        LEFT JOIN empleado_info_O e ON a.id_persona = e.id_persona AND e.idEmpresa = a.idEmpresa
        WHERE a.idEmpresa = %s
    """
    params = [company_id]
    
    if fecha_inicio:
        query += " AND a.fecha >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND a.fecha <= %s"
        params.append(fecha_fin)
    if id_persona:
        query += " AND a.id_persona = %s"
        params.append(id_persona)
    
    query += " ORDER BY a.fecha DESC, a.hora_entrada DESC LIMIT 1000"
    
    cursor.execute(query, params)
    asistencias = cursor.fetchall()
    cursor.close()
    
    if not asistencias:
        raise HTTPException(404, "No data found")
    
    if format == "pdf":
        return generate_asistencia_pdf(asistencias, fecha_inicio, fecha_fin)
    else:
        return generate_asistencia_excel(asistencias, fecha_inicio, fecha_fin)

def generate_asistencia_pdf(asistencias, fecha_inicio, fecha_fin):
    """Generate PDF report for asistencia"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1f2937'), spaceAfter=30, alignment=TA_CENTER)
    elements.append(Paragraph("Reporte de Asistencia", title_style))
    
    if fecha_inicio or fecha_fin:
        date_text = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Fin'}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    total_asistencias = len(asistencias)
    summary = f"Total de registros: {total_asistencias}"
    elements.append(Paragraph(summary, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    data = [['Fecha', 'Empleado', 'CI', 'Cargo', 'Entrada', 'Salida']]
    for a in asistencias:
        empleado = f"{a['nombres_persona'] or ''} {a['apellido_paternoPersona'] or ''}".strip() or 'N/A'
        entrada = str(a['hora_entrada']) if a['hora_entrada'] else '-'
        salida = str(a['hora_salida']) if a['hora_salida'] else '-'
        data.append([
            a['fecha'].strftime('%d/%m/%Y') if a['fecha'] else '-',
            empleado[:20],
            a['ci_persona'] or '-',
            (a['cargo'] or '-')[:15],
            entrada[:8],
            salida[:8]
        ])
    
    table = Table(data, colWidths=[1*inch, 1.5*inch, 1*inch, 1.2*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=asistencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
    )

def generate_asistencia_excel(asistencias, fecha_inicio, fecha_fin):
    """Generate Excel report for asistencia"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "Reporte de Asistencia"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    if fecha_inicio or fecha_fin:
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Fin'}"
        ws['A2'].alignment = Alignment(horizontal="center")
    
    total_asistencias = len(asistencias)
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Total de registros: {total_asistencias}"
    ws['A3'].alignment = Alignment(horizontal="center")
    ws['A3'].font = Font(bold=True)
    
    headers = ['Fecha', 'Empleado', 'CI', 'Cargo', 'Entrada', 'Salida']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row_num, a in enumerate(asistencias, 6):
        empleado = f"{a['nombres_persona'] or ''} {a['apellido_paternoPersona'] or ''}".strip() or 'N/A'
        entrada = str(a['hora_entrada']) if a['hora_entrada'] else '-'
        salida = str(a['hora_salida']) if a['hora_salida'] else '-'
        
        ws.cell(row=row_num, column=1, value=a['fecha'].strftime('%d/%m/%Y') if a['fecha'] else '-').border = border
        ws.cell(row=row_num, column=2, value=empleado).border = border
        ws.cell(row=row_num, column=3, value=a['ci_persona'] or '-').border = border
        ws.cell(row=row_num, column=4, value=a['cargo'] or '-').border = border
        ws.cell(row=row_num, column=5, value=entrada[:8]).border = border
        ws.cell(row=row_num, column=6, value=salida[:8]).border = border
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=asistencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

@app.get("/health")
async def health():
    return {"status": "healthy", "service": "export_service"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8014)
